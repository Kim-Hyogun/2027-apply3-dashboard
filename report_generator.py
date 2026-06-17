import openpyxl
import os
import sys
import datetime

def get_korean_weekday(date_obj):
    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
    return weekdays[date_obj.weekday()]

def generate_report(day_num):
    print(f"[report_generator] Generating report for Day {day_num}...")
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Paths
    template_path = os.path.join(base_dir, "sample", "2027_apply3_report_1day.xlsx")
    input_dir = os.path.join(base_dir, "input-this_years_support_status")
    this_year_input_path = os.path.join(input_dir, f"2027_apply3-{day_num}day-input.xlsx")
    last_year_report_path = os.path.join(base_dir, "input-last_years_application_status", "2026_apply3-report.xlsx")
    output_dir = os.path.join(base_dir, "output-report")
    output_report_path = os.path.join(output_dir, f"2027_report_{day_num}day.xlsx")
    
    # Check if files exist
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template report not found: {template_path}")
    if not os.path.exists(this_year_input_path):
        raise FileNotFoundError(f"This year's input data not found: {this_year_input_path}")
    if not os.path.exists(last_year_report_path):
        raise FileNotFoundError(f"Last year's report data not found: {last_year_report_path}")
        
    os.makedirs(output_dir, exist_ok=True)
    
    # Date calculations
    start_date = datetime.date(2027, 1, 4)  # 2027학년도 정시 모집 지원기간 시작일
    current_date = start_date + datetime.timedelta(days=day_num - 1)
    month = current_date.month
    day = current_date.day
    weekday = get_korean_weekday(current_date)
    d_day = 17 - day_num
    
    b3_text = f"접수{day_num}일차(D-{d_day}) / {month}월{day}일({weekday}) 17시기준"
    sheet_title = f"{day_num}일차{month}월{day}일({weekday})17시"
    c42_text = f"접수{day_num}일차\n17시 기준"
    
    # Load raw daily input file (2027_apply3-Nday-input.xlsx)
    # Load with data_only=True to get raw values from input
    wb_input = openpyxl.load_workbook(this_year_input_path, data_only=True)
    ws_input = wb_input.active
    
    # Parse input data
    # Key: Department name, Value: (in_quota_applicants, out_of_quota_applicants)
    input_data = {}
    for r in range(6, ws_input.max_row + 1):
        dept_name = ws_input.cell(row=r, column=1).value  # Column A
        if not dept_name:
            continue
        dept_name = str(dept_name).strip()
        # In-quota applicants is in column K (11)
        in_quota = ws_input.cell(row=r, column=11).value
        # Out-of-quota applicants is in column AI (35)
        out_quota = ws_input.cell(row=r, column=35).value
        
        # Default to 0 if None
        in_quota = in_quota if in_quota is not None else 0
        out_quota = out_quota if out_quota is not None else 0
        
        input_data[dept_name] = (in_quota, out_quota)
    
    # Load last year's report (2026_apply3-report.xlsx)
    wb_last = openpyxl.load_workbook(last_year_report_path, data_only=True)
    ws_last = wb_last.active
    
    # Parse last year's data
    # Department name mapping: "물리치료학과" -> "물리치료과"
    name_mapping = {"물리치료학과": "물리치료과"}
    last_year_data = {}
    
    # Find matching row in last year's report
    for r in range(6, ws_last.max_row + 1):
        dept_name = ws_last.cell(row=r, column=3).value  # Column C
        if not dept_name:
            continue
        dept_name = str(dept_name).strip()
        
        # Daily cumulative in-quota: column index 5 + N (F for N=1)
        last_in_quota_n = ws_last.cell(row=r, column=5 + day_num).value
        # Daily cumulative out-of-quota: column index 23 + N (X for N=1)
        last_out_quota_n = ws_last.cell(row=r, column=23 + day_num).value
        # In-quota final: column index 23 (W)
        last_in_quota_final = ws_last.cell(row=r, column=23).value
        # Out-of-quota final: column index 41 (AO)
        last_out_quota_final = ws_last.cell(row=r, column=41).value
        
        # Default to 0 if None
        last_in_quota_n = last_in_quota_n if last_in_quota_n is not None else 0
        last_out_quota_n = last_out_quota_n if last_out_quota_n is not None else 0
        last_in_quota_final = last_in_quota_final if last_in_quota_final is not None else 0
        last_out_quota_final = last_out_quota_final if last_out_quota_final is not None else 0
        
        last_year_data[dept_name] = (
            last_in_quota_n,
            last_out_quota_n,
            last_in_quota_final,
            last_out_quota_final
        )
        
    # Get previous day's report values (T & U)
    prev_day_data = {}
    if day_num > 1:
        prev_report_path = os.path.join(output_dir, f"2027_report_{day_num - 1}day.xlsx")
        if not os.path.exists(prev_report_path):
            raise FileNotFoundError(f"Previous day report not found: {prev_report_path}")
        wb_prev = openpyxl.load_workbook(prev_report_path, data_only=True)
        ws_prev = wb_prev.active
        for r in range(6, 41):
            dept_name = ws_prev.cell(row=r, column=3).value  # Column C
            if not dept_name:
                continue
            dept_name = str(dept_name).strip()
            # F (6) is in-quota applicants of prev day, K (11) is out-of-quota applicants of prev day
            prev_in = ws_prev.cell(row=r, column=6).value
            prev_out = ws_prev.cell(row=r, column=11).value
            prev_in = prev_in if prev_in is not None else 0
            prev_out = prev_out if prev_out is not None else 0
            prev_day_data[dept_name] = (prev_in, prev_out)
    else:
        # Day 1 -> previous day is 0
        pass

    # Load template workbook (data_only=False to keep formulas)
    wb_template = openpyxl.load_workbook(template_path, data_only=False)
    ws_template = wb_template.active
    
    # Rename sheet
    ws_template.title = sheet_title
    
    # Update B3 header cell
    ws_template['B3'] = b3_text
    
    # Update C42 (전년도 비교 날짜 기준)
    ws_template['C42'] = c42_text
    
    # Fill in department rows (6 to 40)
    for r in range(6, 41):
        dept_name = ws_template.cell(row=r, column=3).value  # Column C
        if not dept_name:
            continue
        dept_name = str(dept_name).strip()
        
        # 1. Today's cumulative values (F & K)
        in_quota_val, out_quota_val = input_data.get(dept_name, (0, 0))
        ws_template.cell(row=r, column=6).value = in_quota_val  # Column F (정원내 지원인원)
        ws_template.cell(row=r, column=11).value = out_quota_val  # Column K (정원외 지원인원)
        
        # 2. Previous day values (T & U)
        if day_num == 1:
            ws_template.cell(row=r, column=20).value = 0  # Column T
            ws_template.cell(row=r, column=21).value = 0  # Column U
        else:
            prev_in, prev_out = prev_day_data.get(dept_name, (0, 0))
            ws_template.cell(row=r, column=20).value = prev_in   # Column T
            ws_template.cell(row=r, column=21).value = prev_out  # Column U
            
        # 3. Last year's matching department
        last_dept_key = name_mapping.get(dept_name, dept_name)
        if last_dept_key in last_year_data:
            last_in_quota_n, last_out_quota_n, last_in_quota_final, last_out_quota_final = last_year_data[last_dept_key]
            ws_template.cell(row=r, column=22).value = last_in_quota_n       # Column V (전년 정원내 지원인원)
            ws_template.cell(row=r, column=23).value = last_out_quota_n      # Column W (전년 정원외 지원인원)
            ws_template.cell(row=r, column=10).value = last_in_quota_final   # Column J (정원내 전년최종 지원인원)
            ws_template.cell(row=r, column=14).value = last_out_quota_final  # Column N (정원외 전년최종 지원인원)
        else:
            print(f"[report_generator] Warning: Department '{last_dept_key}' not found in last year's report!")
            
        # 4. Overwrite R column formula: =SUM(J{r}+N{r})
        ws_template.cell(row=r, column=18).value = f"=SUM(J{r}+N{r})"    # Column R
        
    # Save the file
    wb_template.save(output_report_path)
    print(f"[report_generator] Report successfully saved: {output_report_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python report_generator.py <day_num>")
        sys.exit(1)
    try:
        day = int(sys.argv[1])
        generate_report(day)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
