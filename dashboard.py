import streamlit as st
import openpyxl
import os
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# Set page config
st.set_page_config(
    page_title="2027학년도 정시 모집 지원현황 실시간 대시보드",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Slate Theme CSS Injection
st.markdown("""
<style>
    /* Dark Slate Palette */
    :root {
        --background-color: #0F172A;
        --card-bg: #1E293B;
        --primary-color: #38BDF8;
        --text-color: #F8FAFC;
        --accent-green: #34D399;
        --accent-gold: #FBBF24;
        --accent-red: #F87171;
    }
    
    .reportview-container {
        background-color: var(--background-color);
        color: var(--text-color);
    }
    
    /* Sleek card styling */
    div[data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 700 !important;
        color: var(--primary-color) !important;
    }
    
    div[data-testid="stMetricDelta"] {
        font-weight: 500 !important;
    }
    
    /* Metric Card Wrapper */
    .metric-card {
        background-color: #1E293B;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #334155;
        box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
        transition: transform 0.2s ease, border-color 0.2s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: #475569;
    }
    
    .metric-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: #94A3B8;
        margin-bottom: 0.5rem;
    }
    
    .metric-value {
        font-size: 1.75rem;
        font-weight: 700;
        color: #38BDF8;
    }
    
    .metric-delta {
        font-size: 0.875rem;
        margin-top: 0.25rem;
        font-weight: 500;
    }
    
    .delta-up {
        color: #34D399;
    }
    .delta-down {
        color: #F87171;
    }
    .delta-neutral {
        color: #94A3B8;
    }

    /* Main Title Styling */
    .dashboard-title {
        font-size: 2.25rem;
        font-weight: 800;
        background: linear-gradient(135deg, #38BDF8 0%, #34D399 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.25rem;
    }
    .dashboard-subtitle {
        font-size: 1rem;
        color: #94A3B8;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# 1. Security Login Gate
def check_password():
    """Returns True if the user has entered the correct password."""
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        try:
            # Check secrets fallback
            secret_pw = st.secrets.get("ipsi_password", "ipsi2027")
        except Exception:
            secret_pw = "ipsi2027"
            
        if st.session_state["password"] == secret_pw:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password in session state
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state or not st.session_state["password_correct"]:
        # Center-align using columns
        _, col_center, _ = st.columns([1, 1.5, 1])
        with col_center:
            st.markdown("<div style='text-align: center; margin-top: 100px;'>", unsafe_allow_html=True)
            st.markdown("<h1>🔐 정시 모집 지원현황 시스템</h1>", unsafe_allow_html=True)
            st.text_input(
                "보안 비밀번호를 입력해주세요",
                type="password",
                on_change=password_entered,
                key="password",
                label_visibility="collapsed"
            )
            st.markdown("<p style='text-align: center; color: #94A3B8; margin-top: 0.5rem;'>보안 비밀번호를 입력해주세요</p>", unsafe_allow_html=True)
            if "password_correct" in st.session_state and not st.session_state["password_correct"]:
                st.error("❌ 비밀번호가 올바르지 않습니다.")
            st.markdown("</div>", unsafe_allow_html=True)
        return False
    else:
        # Password is correct
        return True

if not check_password():
    st.stop()

# 2. Dynamic File Detection (No cache decorator on list/detection function)
def get_available_days():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "output-report")
    if not os.path.exists(output_dir):
        return []
    
    days = []
    # Match output reports: 2027_report_Nday.xlsx
    for filename in os.listdir(output_dir):
        if filename.startswith("2027_report_") and filename.endswith("day.xlsx"):
            try:
                day_num = int(filename.split("report_")[1].split("day")[0])
                days.append(day_num)
            except ValueError:
                pass
    return sorted(days)

available_days = get_available_days()

if not available_days:
    st.warning("⚠️ 아직 생성된 일차별 보고서 파일이 없습니다. 파이프라인(run_pipeline.py)을 가동하여 데이터를 먼저 생성해 주세요.")
    st.stop()

# Sidebar for selection
st.sidebar.markdown("### 📅 보고서 일차 선택")
selected_day = st.sidebar.selectbox(
    "조회할 일차를 선택하세요",
    available_days,
    index=len(available_days) - 1,
    format_func=lambda d: f"{d}일차 보고서"
)

# Load data helper (cached for performance based on selected file)
@st.cache_data
def load_day_report_data(day):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, "output-report", f"2027_report_{day}day.xlsx")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Read department rows 6 to 40
    data = []
    for r in range(6, 41):
        row_dict = {
            "NO": ws.cell(row=r, column=2).value,
            "학과": ws.cell(row=r, column=3).value,
            "입학정원": ws.cell(row=r, column=4).value,
            "정원내_모집": ws.cell(row=r, column=5).value,
            "정원내_지원": ws.cell(row=r, column=6).value,
            "정원내_지원율": ws.cell(row=r, column=7).value,
            "정원내_전일대비": ws.cell(row=r, column=8).value,
            "정원내_전년당일대비": ws.cell(row=r, column=9).value,
            "정원내_전년최종": ws.cell(row=r, column=10).value,
            "정원외_지원": ws.cell(row=r, column=11).value,
            "정원외_전일대비": ws.cell(row=r, column=12).value,
            "정원외_전년당일대비": ws.cell(row=r, column=13).value,
            "정원외_전년최종": ws.cell(row=r, column=14).value,
            "계_지원": ws.cell(row=r, column=15).value,
            "계_전일대비": ws.cell(row=r, column=16).value,
            "계_전년당일대비": ws.cell(row=r, column=17).value,
            "계_전년최종": ws.cell(row=r, column=18).value,
            "정원내_전년당일": ws.cell(row=r, column=22).value,  # Column V
            "정원외_전년당일": ws.cell(row=r, column=23).value,  # Column W
        }
        data.append(row_dict)
        
    df = pd.DataFrame(data)
    
    # Read totals row 41 (using calculated fields or cells)
    totals = {
        "입학정원": ws.cell(row=41, column=4).value,
        "정원내_모집": ws.cell(row=41, column=5).value,
        "정원내_지원": ws.cell(row=41, column=6).value,
        "정원내_지원율": ws.cell(row=41, column=7).value,
        "정원내_전일대비": ws.cell(row=41, column=8).value,
        "정원내_전년당일대비": ws.cell(row=41, column=9).value,
        "정원내_전년최종": ws.cell(row=41, column=10).value,
        "정원외_지원": ws.cell(row=41, column=11).value,
        "정원외_전일대비": ws.cell(row=41, column=12).value,
        "정원외_전년당일대비": ws.cell(row=41, column=13).value,
        "정원외_전년최종": ws.cell(row=41, column=14).value,
        "계_지원": ws.cell(row=41, column=15).value,
        "계_전일대비": ws.cell(row=41, column=16).value,
        "계_전년당일대비": ws.cell(row=41, column=17).value,
        "계_전년최종": ws.cell(row=41, column=18).value,
        "정원내_전년당일": ws.cell(row=41, column=22).value,  # Column V
        "정원외_전년당일": ws.cell(row=41, column=23).value,  # Column W
    }
    
    b3_val = ws['B3'].value
    return df, totals, b3_val

df, totals, date_header = load_day_report_data(selected_day)

# Target values from last year's final: in-quota = J41 (8639), out-of-quota = N41 (1418), Total = 10057
target_in_quota_final = totals["정원내_전년최종"] if totals["정원내_전년최종"] is not None else 8639
target_out_quota_final = totals["정원외_전년최종"] if totals["정원외_전년최종"] is not None else 1418
target_total_final = target_in_quota_final + target_out_quota_final  # Guarantees exactly 10,057

# Header
st.markdown(f"<div class='dashboard-title'>2027학년도 신입생 정시 모집 지원현황</div>", unsafe_allow_html=True)
st.markdown(f"<div class='dashboard-subtitle'>📊 {date_header} (실시간 집계 결과)</div>", unsafe_allow_html=True)

# 3. Metric Cards (Summary statistics)
col1, col2, col3, col4 = st.columns(4)

def format_delta_inline(val):
    if val is None:
        return "0", "color: #94A3B8;"
    if val > 0:
        return f"+{val:,}", "color: #34D399;"
    elif val < 0:
        return f"{val:,}", "color: #F87171;"
    return "0", "color: #94A3B8;"

in_delta_val, in_delta_style = format_delta_inline(totals["정원내_전일대비"])
in_year_delta_val, in_year_delta_style = format_delta_inline(totals["정원내_전년당일대비"])

out_delta_val, out_delta_style = format_delta_inline(totals["정원외_전일대비"])
out_year_delta_val, out_year_delta_style = format_delta_inline(totals["정원외_전년당일대비"])

total_delta_val, total_delta_style = format_delta_inline(totals["계_전일대비"])
total_year_delta_val, total_year_delta_style = format_delta_inline(totals["계_전년당일대비"])

# Calculate competition rate
in_quota_rate = totals["정원내_지원율"]
if in_quota_rate is None or in_quota_rate == 0:
    in_quota_rate = (totals["정원내_지원"] / totals["정원내_모집"]) if totals["정원내_모집"] else 0.0

achievement_rate = (totals["계_지원"] / target_total_final * 100) if target_total_final else 0.0

with col1:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">정원내 지원인원 / 경쟁률</div>
        <div class="metric-value">{totals['정원내_지원']:,}명 / {in_quota_rate:.2f}:1</div>
        <div class="metric-delta">
            <span>전일 대비: <span style="{in_delta_style}">{in_delta_val}</span></span>
            <span style="color: #475569; margin: 0 4px;">|</span>
            <span>전년 대비: <span style="{in_year_delta_style}">{in_year_delta_val}</span></span>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">정원외 지원인원</div>
        <div class="metric-value">{totals['정원외_지원']:,}명</div>
        <div class="metric-delta">
            <span>전일 대비: <span style="{out_delta_style}">{out_delta_val}</span></span>
            <span style="color: #475569; margin: 0 4px;">|</span>
            <span>전년 대비: <span style="{out_year_delta_style}">{out_year_delta_val}</span></span>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">총 지원인원 합계</div>
        <div class="metric-value">{totals['계_지원']:,}명</div>
        <div class="metric-delta">
            <span>전일 대비: <span style="{total_delta_style}">{total_delta_val}</span></span>
            <span style="color: #475569; margin: 0 4px;">|</span>
            <span>전년 대비: <span style="{total_year_delta_style}">{total_year_delta_val}</span></span>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col4:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">작년 최종마감 대비 달성율</div>
        <div class="metric-value">{achievement_rate:.2f}%</div>
        <div class="metric-delta delta-neutral">목표 최종인원: {target_total_final:,}명</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# 4. Charts Section (Tabs inside "지원자 유입 추이 및 분석")
st.markdown("### 📊 지원자 유입 추이 및 분석")

# Load daily historical trends (今年 vs 작년)
@st.cache_data
def load_trend_data(current_day):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    last_year_file = os.path.join(base_dir, "input-last_years_application_status", "2026_apply3-report.xlsx")
    
    # 1. Parse Last Year's daily cumulative totals
    wb_last = openpyxl.load_workbook(last_year_file, data_only=True)
    ws_last = wb_last.active
    
    last_year_totals = []
    for d in range(1, 18):
        # column 5 + d is N-day in-quota, column 23 + d is N-day out-of-quota
        in_q = ws_last.cell(row=41, column=5 + d).value
        out_q = ws_last.cell(row=41, column=23 + d).value
        in_q = in_q if in_q is not None else 0
        out_q = out_q if out_q is not None else 0
        last_year_totals.append(in_q + out_q)
        
    # 2. Parse Current Year's daily cumulative totals (up to current_day)
    this_year_totals = []
    days_indexed = []
    for d in range(1, current_day + 1):
        report_path = os.path.join(base_dir, "output-report", f"2027_report_{d}day.xlsx")
        if os.path.exists(report_path):
            wb_this = openpyxl.load_workbook(report_path, data_only=True)
            ws_this = wb_this.active
            # total is in cell O41
            tot = ws_this.cell(row=41, column=15).value
            this_year_totals.append(tot if tot is not None else 0)
            days_indexed.append(d)
            
    return days_indexed, this_year_totals, list(range(1, 18)), last_year_totals

days_this, totals_this, days_last, totals_last = load_trend_data(selected_day)

# 1. Setup figures
fig_trend = go.Figure()
fig_trend.add_trace(go.Scatter(
    x=[f"{d}일차" for d in days_this],
    y=totals_this,
    mode="lines+markers",
    name="2027학년도 (올해 누적)",
    line=dict(color="#38BDF8", width=4),
    marker=dict(size=8)
))
fig_trend.add_trace(go.Scatter(
    x=[f"{d}일차" for d in days_last],
    y=totals_last,
    mode="lines+markers",
    name="2026학년도 (작년 누적)",
    line=dict(color="#94A3B8", width=2, dash="dash"),
    marker=dict(size=6)
))
fig_trend.update_layout(
    template="plotly_dark",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    margin=dict(l=40, r=40, t=20, b=40),
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    xaxis=dict(showgrid=True, gridcolor="#334155"),
    yaxis=dict(showgrid=True, gridcolor="#334155")
)

df_comp = df.copy()
df_comp["경쟁률"] = df_comp.apply(
    lambda row: (row["정원내_지원"] / row["정원내_모집"]) if row["정원내_모집"] > 0 else 0.0,
    axis=1
)
df_top10 = df_comp.sort_values(by="경쟁률", ascending=True).tail(10)

fig_top10 = px.bar(
    df_top10,
    x="경쟁률",
    y="학과",
    orientation="h",
    color="경쟁률",
    color_continuous_scale="Viridis",
    labels={"경쟁률": "경쟁률 (지원/모집)"}
)
fig_top10.update_layout(
    template="plotly_dark",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    margin=dict(l=40, r=40, t=20, b=40),
    coloraxis_showscale=False,
    xaxis=dict(showgrid=True, gridcolor="#334155"),
    yaxis=dict(showgrid=False)
)

df_decreased = df.copy()
df_decreased["지원자_증감"] = df_decreased["정원내_지원"] - df_decreased["정원내_전년당일"]
df_dec_filtered = df_decreased[df_decreased["지원자_증감"] < 0].copy()
df_dec_filtered["감소폭"] = df_dec_filtered["지원자_증감"].abs()
df_dec_top10 = df_dec_filtered.sort_values(by="감소폭", ascending=True).tail(10)

if not df_dec_top10.empty:
    fig_dec = px.bar(
        df_dec_top10,
        x="감소폭",
        y="학과",
        orientation="h",
        labels={"감소폭": "지원자 감소수 (명)"},
        color="감소폭",
        color_continuous_scale="Reds"
    )
    fig_dec.update_layout(
        template="plotly_dark",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=40, r=40, t=20, b=40),
        coloraxis_showscale=False,
        xaxis=dict(showgrid=True, gridcolor="#334155"),
        yaxis=dict(showgrid=False)
    )
else:
    fig_dec = None

if not df_dec_filtered.empty:
    df_dec_display = df_dec_filtered.sort_values(by="지원자_증감", ascending=True)[
        ["학과", "정원내_지원", "정원내_전년당일", "지원자_증감"]
    ].rename(columns={
        "정원내_지원": "올해 지원(명)",
        "정원내_전년당일": "작년 당일(명)",
        "지원자_증감": "증감수(명)"
    }).reset_index(drop=True)
    table_height = min(300, (len(df_dec_display) + 1) * 35 + 10)
else:
    df_dec_display = None
    table_height = 100

# 2. Render tabs
tab1, tab2, tab3 = st.tabs([
    "📈 일차별 누적 지원자 추이",
    "🏆 정원내 경쟁률 상위 TOP 10 학과",
    "⚠️ 전년 대비 정원내 지원율 하락 학과 분석"
])

with tab1:
    st.plotly_chart(fig_trend, use_container_width=True)

with tab2:
    st.plotly_chart(fig_top10, use_container_width=True)

with tab3:
    if df_dec_display is not None:
        col_tab3_1, col_tab3_2 = st.columns([1, 1])
        with col_tab3_1:
            if fig_dec is not None:
                st.plotly_chart(fig_dec, use_container_width=True)
            else:
                st.info("하락한 학과가 없습니다.")
        with col_tab3_2:
            st.dataframe(df_dec_display, height=table_height, use_container_width=True)
    else:
        st.info("정원내 지원인원이 전년 당일 대비 하락한 학과가 없습니다.")

st.markdown("<br><hr>", unsafe_allow_html=True)

# 5. Department Detailed Status Table
st.markdown("### 📋 학과별 상세 지원 현황표")

# Create a clean display dataframe
df_display = df.copy()

# Add competition rate column in display format
df_display["정원내_경쟁률"] = df_display.apply(
    lambda r: f"{r['정원내_지원율']:.2f}:1" if r['정원내_지원율'] is not None else "0.00:1",
    axis=1
)

# Format difference columns for better visualization (+/-)
def format_diff_col(val):
    if val is None:
        return "0"
    return f"+{val}" if val > 0 else f"{val}"

# Select and rename columns
df_table = df_display[[
    "NO", "학과", "입학정원", "정원내_모집", "정원내_지원", "정원내_경쟁률", 
    "정원내_전일대비", "정원내_전년당일대비", "정원내_전년최종",
    "정원외_지원", "정원외_전일대비", "정원외_전년당일대비", "정원외_전년최종",
    "계_지원", "계_전일대비", "계_전년당일대비", "계_전년최종"
]].copy()

# Map headers
df_table.columns = [
    "NO", "학과", "입학정원", "정원내 모집", "정원내 지원", "정원내 경쟁률",
    "정원내 전일대비", "정원내 전년당일대비", "정원내 전년최종",
    "정원외 지원", "정원외 전일대비", "정원외 전년당일대비", "정원외 전년최종",
    "총지원 합계", "총지원 전일대비", "총지원 전년당일대비", "총지원 전년최종"
]

# Create totals row and append it
total_row = pd.DataFrame([{
    "NO": "-",
    "학과": "계 (합계)",
    "입학정원": totals["입학정원"],
    "정원내 모집": totals["정원내_모집"],
    "정원내 지원": totals["정원내_지원"],
    "정원내 경쟁률": f"{totals['정원내_지원율']:.2f}:1" if totals.get('정원내_지원율') is not None else (f"{(totals['정원내_지원']/totals['정원내_모집']):.2f}:1" if totals['정원내_모집'] else "0.00:1"),
    "정원내 전일대비": totals["정원내_전일대비"],
    "정원내 전년당일대비": totals["정원내_전년당일대비"],
    "정원내 전년최종": totals["정원내_전년최종"],
    "정원외 지원": totals["정원외_지원"],
    "정원외 전일대비": totals["정원외_전일대비"],
    "정원외 전년당일대비": totals["정원외_전년당일대비"],
    "정원외 전년최종": totals["정원외_전년최종"],
    "총지원 합계": totals["계_지원"],
    "총지원 전일대비": totals["계_전일대비"],
    "총지원 전년당일대비": totals["계_전년당일대비"],
    "총지원 전년최종": totals["계_전년최종"]
}])

df_final_table = pd.concat([df_table, total_row], ignore_index=True)

# Format integers and diffs for nicer visualization
for col in df_final_table.columns:
    if col in ["NO", "학과", "정원내 경쟁률"]:
        continue
    # Format integers
    df_final_table[col] = df_final_table[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) and not isinstance(x, str) else x)

# Set dynamic height for table to avoid vertical scrollbar
# Length of table is 35 rows + 1 totals row = 36.
# 36 * 35.5 pixels + 10 pixels padding is around 1300 pixels
dynamic_height = (len(df_final_table) + 1) * 35 + 20

st.dataframe(
    df_final_table,
    height=dynamic_height,
    use_container_width=True
)

st.markdown("<br><hr>", unsafe_allow_html=True)

# 6. Report Image Viewer & Download Section
st.markdown("### 🖼️ 보고서 인쇄 서식 (고해상도 이미지)")
base_dir = os.path.dirname(os.path.abspath(__file__))
image_path = os.path.join(base_dir, "output-report", f"2027_report_{selected_day}day.png")
excel_file_path = os.path.join(base_dir, "output-report", f"2027_report_{selected_day}day.xlsx")

col_img, col_btn = st.columns([3, 1])

with col_img:
    if os.path.exists(image_path):
        st.image(image_path, caption=f"2027학년도 정시 모집 {selected_day}일차 인쇄 서식", use_container_width=True)
    else:
        st.info("선택된 일차의 보고서 이미지를 찾을 수 없습니다. 파이프라인에서 이미지를 생성해 주세요.")

with col_btn:
    st.markdown("### 📥 보고서 다운로드")
    
    # Download Excel Button
    if os.path.exists(excel_file_path):
        with open(excel_file_path, "rb") as f:
            st.download_button(
                label="📁 엑셀 원본 파일 (.xlsx) 다운로드",
                data=f,
                file_name=f"2027_report_{selected_day}day.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.button("📁 엑셀 파일 없음", disabled=True, use_container_width=True)
        
    # Download PNG Button
    if os.path.exists(image_path):
        with open(image_path, "rb") as f:
            st.download_button(
                label="🖼️ 고해상도 이미지 (.png) 다운로드",
                data=f,
                file_name=f"2027_report_{selected_day}day.png",
                mime="image/png",
                use_container_width=True
            )
    else:
        st.button("🖼️ 이미지 파일 없음", disabled=True, use_container_width=True)
